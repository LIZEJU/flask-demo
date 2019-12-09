#encoding:utf-8

'''
读取excel 中的文件
将excelo中的两个数据合并到一起
课程， 人数，学习时间
pip install openpyxl
Looking in indexes: http://pypi.douban.com/simple
Collecting openpyxl
  Downloading http://pypi.doubanio.com/packages/f4/5f/fb8706fba43b46716e252fdd3ffdfe801a63a0f4663b80b6f3421d85ab70/openpyxl-3.0.2.tar.gz (172kB)
     |████████████████████████████████| 174kB 1.1MB/s
ERROR: Package 'openpyxl' requires a different Python: 3.5.3 not in '>=3.6,
python -V
Python 3.5.3

'''

from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
import datetime # 可以用来处理时间相关的数据

def combine():
    '''
    该函数可以用来处理原数据文件：
    1. 合并表格并写入的 combine 表中
    2. 保存原数据文件
    '''
    wb = load_workbook(filename='courses.xlsx')
    for sheet in wb:
        print(sheet.title)
        data = wb.get_sheet_by_name(sheet.title)
    data1 = wb.get_sheet_names()
    dic1 = {}
    dic2 = {}
    data = wb.get_sheet_by_name(data1[0])
    for i in range(2,data.max_row+1):
        # dic1.setdefault(data.cell(row=i,column=2).value,{'datetime':''})
        dic1.setdefault(data.cell(row=i, column=2).value, {'datetime': '', })
        dic1[data.cell(row=i,column=2).value]['p_num'] = data.cell(row=i,column=3).value
        dic1[data.cell(row=i,column=2).value]['datetime'] = data.cell(row=i,column=1).value
    for i in dic1.items():
        print(i)
    data = wb.get_sheet_by_name(data1[1])
    for i in range(2, data.max_row + 1):
        dic2.setdefault(data.cell(row=i, column=2).value, {'datetime': '', })
        dic2[data.cell(row=i, column=2).value]['datetime'] = data.cell(row=i, column=1).value
        dic2[data.cell(row=i, column=2).value]['study'] = data.cell(row=i, column=3).value

    print(dic2)
    for i in dic2.items():
        print(i)

    newwb = Workbook()
    newwb.create_sheet(index=0,title='combine1')

    new_sheet = newwb.get_sheet_by_name('combine1')
    new_sheet.cell(row=1,column=1).value = '课程名称'
    new_sheet.cell(row=1,column=2).value = '学习人数'
    new_sheet.cell(row=1,column=3).value = '学习时间'
    new_sheet.cell(row=1,column=4).value = '创建时间'

    irow = 1
    for r in dic1.keys():
        for k in dic1[r].keys():
            irow += 1
            new_sheet.cell(row=irow,column=1).value = r
            new_sheet.cell(row=irow,column=2).value = dic1[r]['p_num']
            new_sheet.cell(row=irow,column=3).value = dic2[r]['study']
            new_sheet.cell(row=irow,column=4).value = dic2[r]['datetime']
    newwb.save(filename='new_course.xlsx')

def split():
    '''
    该函数可以用来分割文件：
    1. 读取 combine 表中的数据
    2. 将数据按时间分割
    3. 写入不同的数据表中
    '''
    # TODO
    wb = load_workbook(filename='new_course.xlsx')

# 执行
if __name__ == '__main__':
    combine()
    split()