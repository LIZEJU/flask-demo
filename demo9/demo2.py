#encoding:utf-8

from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
import datetime # 可以用来处理时间相关的数据

# 读取文件以及的表
wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']

def combine():
    '''
    该函数可以用来处理原数据文件：
    1. 合并表格并写入的 combine 表中
    2. 保存原数据文件
    '''
    combine_sheet = wb.create_sheet(title='combine')
    combine_sheet.append(['创建时间','课程名称','学习人数','学习时间'])

    for stu in students_sheet.values:
        # print(stu)
        # 去掉表头行
        if stu[2] != '学习人数':
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    combine_sheet.append(list(stu)+[time[2]])
                    # print(list(stu))
    wb.save('courses.xlsx')

def split():
    '''
    该函数可以用来分割文件：
    1. 读取 combine 表中的数据
    2. 将数据按时间分割
    3. 写入不同的数据表中
    '''
    # TODO
    combine_sheet = wb.get_sheet_by_name('combine')
    split_name = []

    for item in combine_sheet.values:
        print(item)
        if item[0] != '创建时间':
            split_name.append(item[0].strftime("%Y"))
    print(split_name)
    for name in set(split_name):
        wb_temp = Workbook()

        wb_temp.remove(wb_temp.active)

        ws = wb_temp.create_sheet(title=name)
        ws.append(['创建时间','课程名称','学习人数','学习时间'])
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime("%Y") == name:
                    ws.append(item_by_year)
        wb_temp.save('{}.xlsx'.format(name))


# 执行
if __name__ == '__main__':
    # combine()
    split()