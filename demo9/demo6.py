#encoding:utf-8

from openpyxl import load_workbook, Workbook


def create_xlsx():
    wb = Workbook()

    ws = wb.create_sheet(title='user')


    ws.append(['名字','年龄','性别','工作'])

    data = [
        ('pwd','27','man','it'),
        ('pwd','27','man','it'),
        ('pwd','27','man','it'),
        ('pwd','27','man','it'),
    ]
    for i in data:
        ws.append(i)
    wb.save('user.xlsx')

def get():
    wb = load_workbook(filename='user.xlsx')
    user = wb['user']


        # print(type(i))
    # 创建一个新的sheet
    user_info = wb.create_sheet(title='userinfo')
    user_info.append(['名字','年龄','性别','工作'])
    for i in user.values:
        if i[0] != '名字':
            user_info.append(i)
    wb.save('user.xlsx')
if __name__ == '__main__':
    # create_xlsx()
    get()